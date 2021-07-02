'''
Prepares Reportcard of class 8 by automating Excel
by, Ashutosh Sharma
'''

import openpyxl
import os

def update(fi):
    path = os.getcwd()
    fil = path + '\\'+fi+'.xlsx'
    print('\"Beauty lies in the eyes of the beholder!\"')
    print('[*] Trying to update...Keep your fingers crossed :D')
    wb = openpyxl.load_workbook(fil)
    wb.create_sheet('COCU')
    ws = wb['COCU']
    ws['A1']='ROLL'
    ws['B1']='NAME'
    ws['C1']='DRAW'
    ws['D1']='SUPW'
    ws['E1']='MUSIC'
    ws['F1']='PHED'
    ws['G1']='HIEL'
    ws['H1']='ENEL'
    ws['I1']='ATT'
    ws['J1']='RANK'
    ws['K1']='DRAW'
    ws['L1']='SUPW'
    ws['M1']='MUSIC'
    ws['N1']='PHED'
    ws['O1']='HIEL'
    ws['P1']='ENEL'
    ws['Q1']='ATT'
    ws['R1']='RANK'
    ws['C2'] = 'T1'
    ws['K2'] = 'T2'
    ws['A3']='=CLASS!A3'
    ws['B3']='=CLASS!B3'
    print('[+] Creating COCU... done')
    wb.create_sheet('ATV')
    ws = wb['ATV']
    ws['A1']='ROLL'
    ws['B1']='NAME'
    ws['c1']='PN'
    ws['D1']='SC'
    ws['E1']='INI'
    ws['F1']='RES'
    ws['G1']='NIW'
    ws['H1']='RIW'
    ws['I1']='PAC'
    ws['J1']='RES'
    ws['K1']='SOS'
    ws['L1']='LEAD'
    ws['M1']='TEAM'
    ws['N1']='PN'
    ws['O1']='SC'
    ws['P1']='INI'
    ws['Q1']='RES'
    ws['R1']='NIW'
    ws['S1']='RIW'
    ws['T1']='PAC'
    ws['U1']='RES'
    ws['V1']='SOS'
    ws['W1']='LEAD'
    ws['X1']='TEAM'
    ws['C2'] = 'T1'
    ws['N2'] = 'T2'
    ws['A3']='=CLASS!A3'
    ws['B3']='=CLASS!B3'
    print('[+] Creating ATV... done')
    wb.create_sheet('REMARKS')
    ws = wb['REMARKS']
    ws['A1']='ROLL'
    ws['B1']='NAME'
    ws['C1']='LEADERSHIP'
    ws['D1']='SCHOOL-LEVEL'
    ws['E1']='OTHER-LEVELS'
    ws['F1']='REMARKS'
    ws['A2']='=CLASS!A3'
    ws['B2']='=CLASS!B3'
    print('[+] Creating REMARKS... done')
    wb.create_sheet('1002')
    ws = wb['1002']
    ws['A1']='ROLL'
    ws['B1']='NAME'
    ws['C1']='ENGI'
    ws['D1']='ENGII'
    ws['E1']='HIN'
    ws['F1']='MAT'
    ws['G1']='PHY'
    ws['H1']='CHE'
    ws['I1']='BIO'
    ws['J1']='H/C'
    ws['K1']='GEO'
    ws['L1']='COM'
    ws['M1']='TOTAL'
    ws['N1']='PER'
    ws['O1']='MSC'
    ws['P1']='GK'
    ws['Q1']='ESP'
    ws['R1']='HSP'
    ws['S1']='ENGIII'
    ws['T1']='RANK'
    ws['A2']='=class!A3'
    ws['B2']='=class!B3'
    ws['C2']='=u1!c2+u2!c2+p1!c2+p2!c2+a1!c2+a2!c2'
    ws['M2']='=sum(c2:l2)'
    ws['o2']='=u1!o2+u2!o2+p1!o2+p2!o2+a1!o2+a2!o2'
    print('[+] Creating 1002... done')
    wb.create_sheet('300')
    ws = wb['300']
    ws['A1']='ROLL'
    ws['B1']='NAME'
    ws['C1']='ENGI'
    ws['D1']='ENGII'
    ws['E1']='HIN'
    ws['F1']='MAT'
    ws['G1']='PHY'
    ws['H1']='CHE'
    ws['I1']='BIO'
    ws['J1']='H/C'
    ws['K1']='GEO'
    ws['L1']='COM'
    ws['M1']='TOTAL'
    ws['N1']='PER'
    ws['O1']='MSC'
    ws['P1']='GK'
    ws['Q1']='ESP'
    ws['R1']='HSP'
    ws['S1']='ENGIII'
    ws['T1']='RANK'
    ws['A2']='=class!A3'
    ws['B2']='=class!B3'
    ws['C2']='=t1!c2+t2!c2+1002!c2'
    ws['o2']='=t1!o2+t2!o2+1002!o2'
    print('[+] Creating 300... done')
    ws = wb['REPORTCARD']
    for i in range(29, 37): 
        ws['D'+str(i)] = '=VLOOKUP($I$4,COCU!$A$3:$V$103,CLASS!A'+str(i-24)+')'
        ws['F'+str(i)] = '=VLOOKUP($I$4,COCU!$A$3:$V$103,CLASS!A'+str(i-16)+')'
    print('[+] Linking REPORTCARD with COCU... done')
    for i in range(29, 40): 
        ws['K'+str(i)] = '=VLOOKUP($I$4,ATV!$A$3:$Z$103,CLASS!A'+str(i-24)+')'
        ws['M'+str(i)] = '=VLOOKUP($I$4,ATV!$A$3:$Z$103,CLASS!A'+str(i-13)+')'
    print('[+] Linking REPORTCARD with ATV... done')
    ws['B41']='="1. Record of Membership/Leadership roles exercised in the school:- "&VLOOKUP($I$4,REMARKS!$A$2:$V$92,3)'
    ws['B42']='="2. Achievements in various Competitions, Sports, Games and Cultural Activities held at School Level :- "&VLOOKUP($I$4,REMARKS!$A$2:$V$92,4)'
    ws['B43']='="3. Achievements in various Competitions, Sports, Games and Cultural Activities held at Inter-School/District/State/National Level :-"&VLOOKUP($I$4,REMARKS!$A$2:$V$92,5)'
    ws['B44']='="General Remarks :-"&VLOOKUP($I$4,REMARKS!$A$2:$V$92,6)'
    print('[+] Linking REPORTCARD with REMARKS... done')
    wb.save(fil)
    print('[*] Congrats! Update complete!')
    return

update(input('What is the workbook-name? [without .xlsx]: '))